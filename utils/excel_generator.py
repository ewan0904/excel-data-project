from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from io import BytesIO


def generate_excel_file(product_df):
    # Load excel template
    wb = load_workbook(filename="assets/excel_template.xlsx")
    ws = wb['Liste']

    # Styling
    thick_bottom = Border(bottom=Side(style="medium"))

    # Counter to keep track of the rows
    starting_row = 4
    row_in_edit = 4

    # Fill in the data from the products
    for i, product in product_df.iterrows():

        # 1 A: Position (right-aligned)
        # Now you can write data starting from row 3, for example:
        ws.cell(row=row_in_edit, column=1, value=product["Position"])
        ws.cell(row=row_in_edit, column=1,).alignment = Alignment(horizontal="right")

        # 2 B: 2. Position (left-aligned)
        ws.cell(row=row_in_edit, column=2, value=product["2. Position"])
        ws.cell(row=row_in_edit, column=2).alignment = Alignment(horizontal="left")

        # 4 D: Menge (mid-aligned)
        ws.cell(row=row_in_edit, column=4, value=product['Menge'])
        ws.cell(row=row_in_edit, column=4).alignment = Alignment(horizontal="center")

        # 5 E: Artikelbeschreibung (left-aligned)
        ws.cell(row=row_in_edit, column=5, value=product['Titel'])
        ws.cell(row=row_in_edit, column=5).alignment = Alignment(horizontal="left")

        # 6 F: Breite (mid-aligned)
        ws.cell(row=row_in_edit, column=6, value=product['Breite'])
        ws.cell(row=row_in_edit, column=6).alignment = Alignment(horizontal="center")

        # 7 G: Tiefe (mid-aligned)
        ws.cell(row=row_in_edit, column=7, value=product['Tiefe'])
        ws.cell(row=row_in_edit, column=7).alignment = Alignment(horizontal="center")

        # 8 H: Höhe (mid-aligned)
        ws.cell(row=row_in_edit, column=8, value=product['Höhe'])
        ws.cell(row=row_in_edit, column=8).alignment = Alignment(horizontal="center")

        # 10 J: Hersteller (mid-aligned)
        ws.cell(row=row_in_edit, column=10, value=product['Hersteller'])
        ws.cell(row=row_in_edit, column=10).alignment = Alignment(horizontal="center")

        # 11 K: Art_Nr (mid-aligned)
        ws.cell(row=row_in_edit, column=11, value=product['Art_Nr'])
        ws.cell(row=row_in_edit, column=11).alignment = Alignment(horizontal="center")

        # 12 L: Preis (right-aligned, price 2 decimal €)
        ws.cell(row=row_in_edit, column=12, value=product['Preis'])
        ws.cell(row=row_in_edit, column=12).alignment = Alignment(horizontal="right")

        # 13 M: 0.0%
        ws.cell(row=row_in_edit, column=13, value=0.0)

        # 14 N: Formula: 12*(1-13)
        ws[f'N{row_in_edit}'] = f'=L{row_in_edit}*(1-M{row_in_edit})'

        # 15 O: Formula: 14 * 4
        ws[f'O{row_in_edit}'] = f'=N{row_in_edit}*D{row_in_edit}'

        # 16 P: Formula: 19-14
        ws[f'P{row_in_edit}'] = f'=S{row_in_edit}-N{row_in_edit}'

        # 17 Q: Formula: 20-15
        ws[f'Q{row_in_edit}'] = f'=T{row_in_edit}-O{row_in_edit}'

        # 18 R: 0.0%
        ws.cell(row=row_in_edit, column=18, value=0.0)

        # 19 S: Formula: 14*(1+18)
        ws[f'S{row_in_edit}'] = f'=N{row_in_edit}*(1+R{row_in_edit})'

        # 20 T: Formula: 19*4
        ws[f'T{row_in_edit}'] = f'=S{row_in_edit}*D{row_in_edit}'

        # 21 U: Preis
        ws.cell(row=row_in_edit, column=21, value=product['Preis'])

        # 22 V: Formula: 21*4
        ws[f'V{row_in_edit}'] = f'=U{row_in_edit}*D{row_in_edit}'

        # 23 W: Url
        ws.cell(row=row_in_edit, column=23, value=product['Url'])

        # Increase the row_in_edit
        row_in_edit +=1


    # Fill in aggregates and other information
    # --- Jump two rows ---
    row_2 = row_in_edit + 2

    # Zwischensumme UVP Einkauf, Formula: SUM(L_starting_row:L_row_in_edit)
    ws[f'L{row_2}'] = f'=SUM(L{starting_row}:L{row_in_edit})'

    # Zwischensumme TL_EK Einkauf, Formula: SUM(O_starting_row:O_row_in_edit)
    ws[f'O{row_2}'] = f'=SUM(O{starting_row}:O{row_in_edit})'

    # Zwischensumme TL_Marge Marge, Formula: SUM(Q_starting_row:Q_row_in_edit)
    ws[f'Q{row_2}'] = f'=SUM(Q{starting_row}:Q{row_in_edit})'

    # Zwischensumme TL VK Kunde Abgerund, Formula: SUM(V_starting_row:V_row_in_edit)
    ws[f'V{row_2}'] = f'=SUM(V{starting_row}:V{row_in_edit})'

    # 23 W: "Zwischensumme" (bold)
    ws.cell(row=row_2, column=23, value='Zwischensumme')
    ws.cell(row=row_2, column=23).font = Font(bold=True)

    # --- Jump three rows ---
    row_3 = row_in_edit + 3

    # 19 T: "Rabatt" (right-aligned)
    ws.cell(row=row_3, column=19, value='Rabatt')
    ws.cell(row=row_3, column=19).alignment = Alignment(horizontal="right")


    # 20 U: 0.0 (left-aligned)
    ws.cell(row=row_3, column=20, value=0.0)
    ws.cell(row=row_3, column=20).number_format = '0.00'
    ws.cell(row=row_3, column=20).alignment = Alignment(horizontal="left")

    # 22 V: Formula: (22_row_in_edit+2)/100 * 20_row_in_edit+3 (Border bottom)
    ws[f'V{row_3}'] = f'=V{row_2}/100*T{row_3}'
    ws[f'V{row_3}'].border = thick_bottom

    # 23 W: "Rabatt Satz" (bold, left-aligned)
    ws.cell(row=row_3, column=23, value="Rabatt Satz")
    ws.cell(row=row_3, column=23).font = Font(bold=True)

    # --- Jump four rows ---
    row_4 = row_in_edit + 4

    # 21 U: "Netto" (bold, right-aligned)
    ws.cell(row=row_4, column=21, value="Netto")
    ws.cell(row=row_4, column=21).alignment = Alignment(horizontal="right")
    ws.cell(row=row_4, column=21).font = Font(bold=True)

    # 22 V: Formula: 22_row_in_edit+2 - 22_row_in_edit+3 (price 2 decimal €)
    ws[f'V{row_4}'] = f'=V{row_2}-V{row_3}'

    # --- Jump five rows ---
    row_5 = row_in_edit + 5

    # 21 U: "MwSt" (bold, right-aligned)
    ws.cell(row=row_5, column=21, value="MwSt")
    ws.cell(row=row_5, column=21).alignment = Alignment(horizontal="right")
    ws.cell(row=row_5, column=21).font = Font(bold=True)

    # 22 V: Formula: (22_row_in_edit+4 / 100) * 19 (price 2 decimal €, border bottom)
    ws[f'V{row_5}'] = f'=(V{row_4}/100)*19'
    ws[f'V{row_5}'].border = thick_bottom

    # --- Jump six rows ---
    row_6 = row_in_edit + 6

    # 21 U: "Brutto" (bold, right-aligned)
    ws.cell(row=row_6, column=21, value='Brutto')
    ws.cell(row=row_6, column=21).alignment = Alignment(horizontal="right")
    ws.cell(row=row_6, column=21).font = Font(bold=True)

    # 22 V: Formula: 22_row_in_edit+4 + 22_row_in_edit+5 (price 2 decimal €, border bottom)
    ws[f'V{row_6}'] = f'=V{row_4}+V{row_5}'
    ws[f'V{row_6}'].border = thick_bottom

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output